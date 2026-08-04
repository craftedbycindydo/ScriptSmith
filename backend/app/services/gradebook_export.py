"""Gradebook workbook builder.

Kept separate from the router so the same sheet-building code can be exercised
outside a request (scripts, fixtures) without duplicating the layout.
"""

import io
from typing import Any, Dict, List, Optional, Tuple


def _status_of(cell: Optional[Any]) -> Optional[str]:
    """None for a missing submission - the grey fill carries that meaning."""
    if cell is None:
        return None
    status = cell["status"] if isinstance(cell, dict) else cell.status
    return "success" if status == "success" else "error"


def _get(obj: Any, key: str):
    return obj.get(key) if isinstance(obj, dict) else getattr(obj, key, None)


def build_gradebook_workbook(
    classroom_name: str,
    students: List[Any],
    templates: List[Any],
    cells: List[Any],
) -> io.BytesIO:
    """Two sheets, both one row per student.

    Matrix  - one column per template, holding the status.
    Details - each template split into a status column and an output column.

    Accepts either pydantic models or plain dicts so callers outside the API
    (fixtures, scripts) can pass decoded JSON directly.
    """
    from openpyxl import Workbook

    by_cell: Dict[Tuple[int, int], Any] = {
        (_get(c, "user_id"), _get(c, "template_id")): c for c in cells
    }

    wb = Workbook()

    matrix = wb.active
    matrix.title = "Matrix"
    matrix.append(
        ["Student", "Username", "Email", "Success"]
        + [_get(t, "name") for t in templates]
    )
    for s in students:
        statuses = [
            _status_of(by_cell.get((_get(s, "user_id"), _get(t, "template_id"))))
            for t in templates
        ]
        ok = sum(1 for v in statuses if v == "success")
        matrix.append(
            [
                _get(s, "name"),
                _get(s, "username"),
                _get(s, "email") or "",
                f"{ok}/{len(templates)}",
            ]
            + statuses
        )
    matrix.append(
        ["Submitted", "", "", ""]
        + [_get(t, "submitted_count") for t in templates]
    )

    # One column per template, like Matrix, but the cell carries the raw output.
    # The status is conveyed by the fill colour, so it is not repeated as text.
    details = wb.create_sheet("Details")
    details.append(
        ["Student", "Username", "Email", "Success"]
        + [_get(t, "name") for t in templates]
    )
    state_grid: List[List[Optional[str]]] = []
    for s in students:
        values: List[Any] = []
        statuses: List[Optional[str]] = []
        for t in templates:
            cell = by_cell.get((_get(s, "user_id"), _get(t, "template_id")))
            state = _status_of(cell)
            statuses.append(state)
            text = ""
            if cell is not None:
                text = _get(cell, "error_message") or _get(cell, "output") or ""
            values.append(text or None)
        state_grid.append(statuses)
        ok = sum(1 for v in statuses if v == "success")
        details.append(
            [
                _get(s, "name"),
                _get(s, "username"),
                _get(s, "email") or "",
                f"{ok}/{len(templates)}",
            ]
            + values
        )

    # Keep the identifying columns visible while template columns scroll.
    for sheet in (matrix, details):
        sheet.freeze_panes = "E2"
        sheet.column_dimensions["A"].width = 26
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 26

    _style(matrix, len(templates), first_data_col=5, states=state_grid)
    _style(
        details,
        len(templates),
        first_data_col=5,
        states=state_grid,
        wrap=True,
        body_width=38,
        row_height=64,
    )

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def workbook_filename(classroom_name: str) -> str:
    safe = "".join(ch if ch.isalnum() else "_" for ch in classroom_name)
    return f"{safe}_gradebook.xlsx"


# Excel's own Good/Bad/Neutral palette, so the sheet reads natively.
_FILL = {
    "success": ("C6EFCE", "006100"),
    "error": ("FFC7CE", "9C0006"),
    None: ("F2F2F2", "808080"),
}


def _style(
    sheet,
    n_templates: int,
    first_data_col: int,
    states: List[List[Optional[str]]],
    wrap: bool = False,
    body_width: Optional[int] = None,
    row_height: int = 18,
) -> None:
    """Title row, gridlines on every cell, status-coloured template cells."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    title_fill = PatternFill("solid", fgColor="44546A")
    title_font = Font(bold=True, color="FFFFFF", size=11)
    hair = Side(style="thin", color="BFBFBF")
    box = Border(left=hair, right=hair, top=hair, bottom=hair)

    last_col = first_data_col + n_templates - 1

    for cell in sheet[1]:
        cell.fill = title_fill
        cell.font = title_font
        # No wrap: wrapping narrow columns shreds names like
        # "Homework Assignment 1" into "k Assignme nt 1".
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = box
    sheet.row_dimensions[1].height = 22

    # Width follows the template name so headers stay readable, unless the
    # sheet needs a fixed width to contain wrapped body text.
    for idx in range(first_data_col, last_col + 1):
        name = sheet.cell(1, idx).value or ""
        width = body_width or max(12, min(len(str(name)) + 3, 34))
        sheet.column_dimensions[get_column_letter(idx)].width = width

    # wrap_text keeps long output inside its own cell. Without it Excel spills
    # text across neighbouring empty cells, which is unreadable in a grid.
    body = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    for row in sheet.iter_rows(min_row=2, max_col=last_col):
        for cell in row:
            cell.border = box
            cell.alignment = body
        sheet.row_dimensions[row[0].row].height = row_height

    # Fills come from the status grid, not the cell text: Details cells hold
    # only the output, so the colour is what tells success from error.
    for r, row_states in enumerate(states):
        for c, state in enumerate(row_states):
            bg, fg = _FILL[state]
            cell = sheet.cell(row=r + 2, column=first_data_col + c)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(bold=True, color=fg)
